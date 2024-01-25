
from openpyxl import load_workbook
# Loading Workbook in read mode
wb = load_workbook(filename='Book.xlsx',read_only=True)
# getting work
ws = wb['Sheet1']
#for row in ws.rows:
  #print(row)
for x in range(1,5):
  if (ws.cell(row=x,column=1).value)=='0x53ae22ad':
    print(ws.cell(row=x,column=2).value)
    print(ws.cell(row=x,column=3).value)
    print(ws.cell(row=x,column=4).value)
  #print(type(ws.cell(row=x,column=1).value))



#for row in ws.rows:
  #for cell in row:
    #print(cell.value)

# close the workbook after reading
wb.close()

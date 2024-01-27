import paho.mqtt.client as mqtt
from datetime import datetime
from openpyxl import load_workbook

# Loading Workbook in read mode
wb = load_workbook(filename='Book.xlsx',read_only=True)
# getting work
ws = wb['Sheet1']

def GetUidDetail(uidnumber):
    for x in range(1,5):
        if (ws.cell(row=x,column=1).value)== uidnumber:
            print(ws.cell(row=x,column=2).value)
            file.write(ws.cell(row=x,column=2).value+",")
            print(ws.cell(row=x,column=3).value)
            file.write(ws.cell(row=x,column=3).value+",")
            print(ws.cell(row=x,column=4).value)
            file.write(ws.cell(row=x,column=4).value+"\n")
MQTT_ADDRESS = "10.3.141.1"
MQTT_USER = "npdAtom"
MQTT_PASSWORD = "npd@Atom"
MQTT_TOPIC = 'RFID DATA'

# open the csv file
file = open("log.csv","a")

def on_connect(client,userdata,flags,rc):
    print('Connected with result code ' + str(rc))
    print('')
    client.subscribe(MQTT_TOPIC)

def on_message(client,userdata,msg):
    # Loading Workbook in read mode
    #wb = load_workbook(filename='Book.xlsx',read_only=True)
    # getting work
    #ws = wb['Sheet1']
    #print(type(str(msg.payload)[7:17]))
    print('..................................')
    #print(msg.topic + ' ' + str(msg.payload))
    #print(msg.topic)
    #file.write(str(msg.topic)+",")
    #print(mqtt_client)
    print(str(msg.payload)[2:17])
    file.write(str(msg.payload)[7:17]+",")
    print(str(msg.payload)[18:22])
    file.write(str(msg.payload)[18:22]+",")
    print('TimeStamp:', datetime.now())
    file.write(datetime.today().strftime('%Y-%m-%d'+","+'%H:%M:%S')+",")
    GetUidDetail(str(msg.payload)[7:17])
    #wb.close()

def main():
    mqtt_client = mqtt.Client()
    #print('mqtt_client')
    mqtt_client.username_pw_set(MQTT_USER, MQTT_PASSWORD)
    mqtt_client.on_connect = on_connect
    mqtt_client.on_message = on_message

    mqtt_client.connect(MQTT_ADDRESS,1883)
    mqtt_client.loop_forever()

if __name__ == '__main__':
    print('MQTT to InfluxDB bridge')
    main()
